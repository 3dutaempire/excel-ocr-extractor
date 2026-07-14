import streamlit as st
import os
import io
import re
import zipfile
import rarfile
import datetime
import tempfile
import shutil

import openpyxl
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.styles import Font
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract

# ==========================================
# KONFIGURASI HALAMAN
# ==========================================
st.set_page_config(
    page_title="Excel Image & Email Extractor",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==========================================
# HELPER: FORMAT TANGGAL BAHASA INDONESIA
# ==========================================
INDONESIAN_MONTHS = {
    1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
    5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
}

def format_date_indonesian(dt):
    day = dt.strftime('%d')
    month = INDONESIAN_MONTHS[dt.month]
    year = dt.strftime('%Y')
    return f"{day} {month} {year}"

# ==========================================
# KONFIGURASI TESSERACT
# ==========================================
def check_tesseract():
    """Cek apakah Tesseract tersedia"""
    try:
        pytesseract.get_tesseract_version()
        return True
    except Exception as e:
        st.error(f"Tesseract OCR tidak ditemukan: {e}")
        st.info("Pastikan Tesseract sudah terinstall di server.")
        return False

# ==========================================
# FUNGSI PEMROSESAN GAMBAR & OCR
# ==========================================
def preprocess_image(image_stream):
    """Preprocessing gambar untuk OCR yang lebih baik"""
    img = Image.open(image_stream)
    img = img.convert('L')  # Convert to grayscale
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    img = img.filter(ImageFilter.SHARPEN)
    return img

def extract_email_from_bytes(image_bytes):
    """Ekstrak email dari gambar menggunakan OCR"""
    try:
        processed = preprocess_image(io.BytesIO(image_bytes))
        text = pytesseract.image_to_string(processed, lang='eng')
        pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        emails = re.findall(pattern, text)
        return emails[0] if emails else "Email tidak ditemukan"
    except Exception as e:
        return f"Error: {str(e)}"

# ==========================================
# FUNGSI SETUP EXCEL
# ==========================================
def setup_sheet(ws):
    """Setup header dan dimensi kolom Excel"""
    headers = ["No", "Nama", "Date", "Email", "Screen Shoot"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 30
    ws.row_dimensions[1].height = 20

# ==========================================
# FUNGSI PROSES SINGLE IMAGE
# ==========================================
def process_single_image(ws, row, item):
    """Proses satu gambar dan masukkan ke sheet Excel"""
    FIXED_ROW_PT = 100
    FIXED_ROW_PX = int(FIXED_ROW_PT * 1.333)
    
    pil_img = Image.open(io.BytesIO(item['bytes']))
    ow, oh = pil_img.size
    ratio = FIXED_ROW_PX / oh
    nw = int(ow * ratio)
    nh = FIXED_ROW_PX
    
    ws.cell(row=row, column=1, value=row - 1)
    ws.cell(row=row, column=2, value=item['name'])
    ws.cell(row=row, column=3, value=item['date'])
    
    email = extract_email_from_bytes(item['bytes'])
    ws.cell(row=row, column=4, value=email)
    
    exc_img = OpenPyXLImage(io.BytesIO(item['bytes']))
    exc_img.width = nw
    exc_img.height = nh
    ws.add_image(exc_img, f'E{row}')
    ws.row_dimensions[row].height = FIXED_ROW_PT
    
    return email

# ==========================================
# FUNGSI EKSTRAK FILE DARI UPLOAD
# ==========================================
def extract_files_from_uploads(uploaded_files, progress_bar, status_text):
    """Ekstrak gambar dari file yang diupload (ZIP/RAR atau gambar langsung)"""
    exts = ('.jpg', '.jpeg', '.png', '.bmp', '.gif')
    groups = []
    total_files = len(uploaded_files)
    
    for idx, uploaded_file in enumerate(uploaded_files):
        status_text.text(f"📂 Memproses file {idx + 1}/{total_files}: {uploaded_file.name}")
        progress_bar.progress((idx / total_files) * 100)
        
        file_ext = os.path.splitext(uploaded_file.name.lower())[1]
        items = []
        
        try:
            if file_ext == '.zip':
                # Proses file ZIP
                with zipfile.ZipFile(uploaded_file, 'r') as z:
                    for entry in sorted(z.namelist()):
                        if entry.lower().endswith(exts) and not entry.startswith('__MACOSX'):
                            try:
                                info = z.getinfo(entry)
                                dt = datetime.datetime(*info.date_time)
                                items.append({
                                    'name': entry,
                                    'bytes': z.read(entry),
                                    'date': format_date_indonesian(dt)
                                })
                            except:
                                pass
                                
            else:
                # File gambar langsung
                if file_ext in exts:
                    file_bytes = uploaded_file.getvalue()
                    dt = datetime.datetime.now()
                    items.append({
                        'name': uploaded_file.name,
                        'bytes': file_bytes,
                        'date': format_date_indonesian(dt)
                    })
            
            if items:
                base_name = os.path.splitext(uploaded_file.name)[0]
                groups.append({'name': base_name, 'items': items})
                
        except Exception as e:
            st.error(f"❌ Gagal memproses {uploaded_file.name}: {e}")
            continue
    
    return groups

# ==========================================
# FUNGSI UTAMA PROSES
# ==========================================
def process_images_to_excel(groups, separate_files, progress_bar, status_text, log_area):
    """Proses semua gambar menjadi Excel"""
    total_items = sum(len(g['items']) for g in groups)
    
    if total_items == 0:
        st.error("❌ Tidak ada file gambar yang ditemukan!")
        return None, None
    
    log_area.text(f"✅ Ditemukan total {total_items} gambar. Memulai OCR...")
    
    global_idx = 0
    errors = 0
    excel_files = []
    
    # Buat workbook gabungan jika diperlukan
    if not separate_files:
        wb_combined = openpyxl.Workbook()
        ws_combined = wb_combined.active
        setup_sheet(ws_combined)
    
    for group_idx, group in enumerate(groups):
        status_text.text(f" Memproses grup: {group['name']}")
        log_area.text(f"📂 Memproses grup: {group['name']} ({len(group['items'])} gambar)")
        
        if separate_files:
            # Buat workbook terpisah untuk setiap grup
            wb = openpyxl.Workbook()
            ws = wb.active
            setup_sheet(ws)
            current_row = 2
        else:
            # Gunakan workbook gabungan
            ws = ws_combined
            current_row = global_idx + 2
        
        for item_idx, item in enumerate(group['items']):
            try:
                email = process_single_image(ws, current_row, item)
                if email.startswith("Error:"):
                    errors += 1
                    log_area.text(f"  ⚠️ Warning: {item['name']} - {email}")
            except Exception as e:
                log_area.text(f"  ❌ Error proses {item['name']}: {e}")
                ws.cell(row=current_row, column=4, value=f"Error: {e}")
                errors += 1
            
            current_row += 1
            global_idx += 1
            
            # Update progress
            progress = (global_idx / total_items) * 100
            progress_bar.progress(progress)
            
            if global_idx % 10 == 0:
                status_text.text(f"⏳ Memproses gambar {global_idx}/{total_items}...")
        
        # Simpan file terpisah jika mode separate
        if separate_files:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            excel_files.append((f"{group['name']}.xlsx", output.getvalue()))
            log_area.text(f"  ✅ Grup {group['name']} selesai")
    
    # Simpan file gabungan
    if not separate_files:
        output = io.BytesIO()
        wb_combined.save(output)
        output.seek(0)
        excel_files.append(("Combined_Output.xlsx", output.getvalue()))
        log_area.text(f"✅ File Excel gabungan selesai dibuat!")
    
    if errors > 0:
        log_area.text(f"⚠️ Total error/warning: {errors}")
    
    log_area.text("🎉 Semua proses selesai!")
    return excel_files, errors

# ==========================================
# UI STREAMLIT
# ==========================================
def main():
    st.title("📊 Excel Image & Email Extractor")
    st.markdown("---")
    
    # Sidebar untuk konfigurasi
    with st.sidebar:
        st.header("️ Pengaturan")
        
        # Cek Tesseract
        if not check_tesseract():
            st.stop()
        
        separate_mode = st.checkbox(
            "📁 Pisahkan file Excel per ZIP/RAR",
            value=False,
            help="Aktifkan untuk membuat file Excel terpisah untuk setiap file ZIP/RAR yang diupload"
        )
        
        st.markdown("---")
        st.info("**Format yang didukung:**\n- ZIP\n- RAR\n- JPG, JPEG, PNG, BMP, GIF")
        
        st.markdown("---")
        st.markdown("**Cara penggunaan:**\n1. Upload file ZIP/RAR atau gambar\n2. Tunggu proses OCR\n3. Download hasil Excel")
    
    # Area upload file
    st.subheader("📤 Upload File")
    uploaded_files = st.file_uploader(
        "Pilih file ZIP, RAR, atau gambar (bisa multiple)",
        type=['zip', 'jpg', 'jpeg', 'png', 'bmp', 'gif'],
        accept_multiple_files=True
    )
    
    if uploaded_files:
        st.write(f"📋 **{len(uploaded_files)} file dipilih:**")
        for f in uploaded_files:
            size_kb = f.size / 1024
            st.text(f"  • {f.name} ({size_kb:.1f} KB)")
        
        st.markdown("---")
        
        # Tombol mulai proses
        col1, col2 = st.columns([3, 1])
        with col1:
            start_btn = st.button("🚀 Mulai Proses OCR", type="primary", use_container_width=True)
        with col2:
            clear_btn = st.button("🗑️ Reset", use_container_width=True)
        
        if clear_btn:
            st.cache_data.clear()
            st.rerun()
        
        if start_btn:
            # Progress bar dan status
            progress_bar = st.progress(0)
            status_text = st.empty()
            log_area = st.empty()
            
            # Area log
            with st.expander("📝 Detail Log Proses", expanded=True):
                log_text = st.empty()
            
            status_text.text("⏳ Mempersiapkan ekstraksi file...")
            
            # Ekstrak file dari upload
            groups = extract_files_from_uploads(uploaded_files, progress_bar, status_text)
            
            if groups:
                # Proses ke Excel
                excel_files, errors = process_images_to_excel(
                    groups, 
                    separate_mode, 
                    progress_bar, 
                    status_text, 
                    log_text
                )
                
                if excel_files:
                    st.markdown("---")
                    st.success("✅ **Proses Selesai!** Download file Excel Anda:")
                    
                    # Tombol download
                    if separate_mode:
                        st.subheader("📥 Download File Excel (Terpisah)")
                        for filename, file_bytes in excel_files:
                            st.download_button(
                                label=f"📄 Download {filename}",
                                data=file_bytes,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                    else:
                        st.subheader(" Download File Excel (Gabungan)")
                        st.download_button(
                            label="📄 Download Combined_Output.xlsx",
                            data=excel_files[0][1],
                            file_name="Combined_Output.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    # Statistik
                    total_images = sum(len(g['items']) for g in groups)
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Gambar", total_images)
                    with col2:
                        st.metric("File Excel", len(excel_files))
                    with col3:
                        st.metric("Error/Warning", errors if errors else 0)
            else:
                st.warning("⚠️ Tidak ada gambar yang bisa diproses.")
    
    else:
        # Tampilan awal
        st.info(" Upload file ZIP, RAR, atau gambar untuk memulai")
        
        st.markdown("### 📖 Fitur Aplikasi:")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            - ✅ **OCR Otomatis** - Ekstrak teks dari gambar
            - ✅ **Deteksi Email** - Cari email dalam gambar
            - ✅ **Format Tanggal Indonesia** - Format yang mudah dibaca
            - ✅ **Multi-File Support** - Upload banyak file sekaligus
            """)
        with col2:
            st.markdown("""
            - ✅ **ZIP & RAR** - Ekstrak otomatis dari archive
            - ✅ **Excel Otomatis** - Generate Excel dengan gambar
            - ✅ **Responsive** - Bisa diakses dari HP
            - ✅ **No Install** - Langsung pakai di browser
            """)

# ==========================================
# ENTRY POINT
# ==========================================
if __name__ == "__main__":
    main()
